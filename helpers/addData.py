from os import name
from flaskext.mysql import MySQL;
import openpyxl
from flask import Flask, request
from openpyxl.xml.constants import ARC_SHARED_STRINGS
from pymysql import connections
from .password import hash_password
app = Flask('__name__')

app.config['MYSQL_DATABASE_HOST'] = 'localhost'
app.config['MYSQL_DATABASE_USER'] = 'root'
app.config['MYSQL_DATABASE_PASSWORD'] = '12345678'
app.config['MYSQL_DATABASE_DB']='dtbank' 
# app.config['MYSQL_DATABASE_PORT']=3306

mysql = MySQL(app)

mysql.init_app(app)
connection=mysql.connect()

cur = connection.cursor()
path = "/Users/ramazanbulut/Documents/GitHub/321_3/helpers/data.xlsx"


def addData():
    cur.execute("CREATE TABLE `dtbank`.`DBManager` (username VARCHAR(45) NOT NULL, password MEDIUMTEXT NOT NULL,PRIMARY KEY (username))")



    cur.execute("""
    CREATE TABLE `dtbank`.`users` (
  `username` VARCHAR(45) NOT NULL,
  `name` VARCHAR(45) NOT NULL,
  `password` MEDIUMTEXT NOT NULL,
  `institution` VARCHAR(45) NOT NULL,
  PRIMARY KEY (`username`, `institution`));

    """)

    cur.execute("CREATE TABLE `dtbank`.`UniProt` (`id` VARCHAR(45) NOT NULL,`seq` MEDIUMTEXT NOT NULL,`name` MEDIUMTEXT NULL,PRIMARY KEY (`id`));")  



    cur.execute("CREATE TABLE `dtbank`.`SideEffect` (`id` VARCHAR(45) NOT NULL,`name` MEDIUMTEXT NOT NULL,PRIMARY KEY (`id`));")



    cur.execute("CREATE TABLE `dtbank`.`Drugs` (`id` VARCHAR(45) NOT NULL,`name` MEDIUMTEXT NOT NULL,`smile` MEDIUMTEXT NULL,`description` MEDIUMTEXT NULL,PRIMARY KEY (`id`));")



    cur.execute("CREATE TABLE `dtbank`.`InteractsWith` (`id1` VARCHAR(45) NOT NULL,`id2` VARCHAR(45) NOT NULL,PRIMARY KEY (`id1`, `id2`),CONSTRAINT `-2` FOREIGN KEY (`id1`)REFERENCES `dtbank`.`Drugs` (`id`)ON DELETE CASCADE ON UPDATE CASCADE,CONSTRAINT `-1`FOREIGN KEY (`id2`)REFERENCES `dtbank`.`Drugs` (`id`)ON DELETE CASCADE ON UPDATE CASCADE);")
    
    cur.execute("""
    CREATE TABLE `dtbank`.`contributors` (
  `doi` VARCHAR(45) NOT NULL,
  `author` VARCHAR(45) NOT NULL,
  `institution` VARCHAR(45) NOT NULL,
   PRIMARY KEY (`doi`, `author`, `institution`),

  CONSTRAINT `key12`
    FOREIGN KEY (`author` , `institution`)
    REFERENCES `dtbank`.`Users` (`username` , `institution`)
    ON DELETE CASCADE
    ON UPDATE CASCADE);
    """)

    cur.execute("""
    CREATE TABLE `dtbank`.`publications` (
  `doi` VARCHAR(45) NOT NULL,
  `institution` VARCHAR(45) NOT NULL,
   PRIMARY KEY (`doi`, `institution`)
  );
    """)
    cur.execute("""
    CREATE TABLE `dtbank`.`institution` (
  `name` VARCHAR(45) NOT NULL,
  `points` INT NOT NULL,
   PRIMARY KEY (`name`)
  );
    """)


    cur.execute("""
    CREATE TABLE `dtbank`.`BindingDB` (
  `drug_id` VARCHAR(45) NOT NULL,
  `prot_id` VARCHAR(45) NOT NULL,
  `reaction_id` VARCHAR(45) NOT NULL,
  `affinity` REAL NULL,
  `measure` VARCHAR(45) NULL,
  `doi` VARCHAR(45) NOT NULL,
  PRIMARY KEY (`reaction_id`),
  INDEX `k1_idx` (`drug_id` ASC) VISIBLE,
  INDEX `k2_idx` (`prot_id` ASC) VISIBLE,
  INDEX `k3_idx` (`doi` ASC) VISIBLE,
  CONSTRAINT `k1`
    FOREIGN KEY (`drug_id`)
    REFERENCES `dtbank`.`Drugs` (`id`)
    ON DELETE CASCADE
    ON UPDATE CASCADE,
  CONSTRAINT `k2`
    FOREIGN KEY (`prot_id`)
    REFERENCES `dtbank`.`UniProt` (`id`)
    ON DELETE CASCADE
    ON UPDATE CASCADE,
  CONSTRAINT `k3`
    FOREIGN KEY (`doi`)
    REFERENCES `dtbank`.`publications` (`doi`)
    ON DELETE CASCADE
    ON UPDATE CASCADE);

    """)
    
    cur.execute("CREATE TABLE `dtbank`.`SIDER` (`drug_id` VARCHAR(45) NOT NULL,`side_id` VARCHAR(45) NOT NULL, PRIMARY KEY (`drug_id`, `side_id`), CONSTRAINT `key3` FOREIGN KEY (`drug_id`) REFERENCES `dtbank`.`Drugs` (`id`) ON DELETE CASCADE ON UPDATE CASCADE, CONSTRAINT `key4` FOREIGN KEY (`side_id`) REFERENCES `dtbank`.`SideEffect` (`id`) ON DELETE CASCADE ON UPDATE CASCADE);")
    
    print('here')
    data = openpyxl.load_workbook(path)

    userSheet=data['User']

    nofRows=userSheet.max_row
    print(nofRows)

    for i in range(2, nofRows + 1):
        name = userSheet.cell(row = i, column = 1).value
        username = userSheet.cell(row = i, column = 2).value
        institute = userSheet.cell(row = i, column = 3).value
        password = userSheet.cell(row = i, column = 4).value
        hashed_password=hash_password(password)
        cur.execute("INSERT INTO `dtbank`.`Users` ( `name`,`username`, `password`, `institution`) VALUES ( %s, %s, %s, %s)",(name,username,hashed_password,institute))    
        connection.commit()
       
    
    DBmanagerSheet=data['Database Manager']

    nofRows=DBmanagerSheet.max_row
    
    for i in range(2, nofRows + 1):
        username = DBmanagerSheet.cell(row = i, column = 1).value
        password = DBmanagerSheet.cell(row = i, column = 2).value
        hashed_password=hash_password(password)
        cur.execute("INSERT INTO `dtbank`.`DBManager` ( `username`, `password`) VALUES ( %s, %s)",(username,hashed_password))    
        connection.commit()
       
     
    Drugbank=data['DrugBank']

    nofRows=Drugbank.max_row
   

    for i in range(2, nofRows + 1):
        id1 = Drugbank.cell(row = i, column = 1).value
        name = Drugbank.cell(row = i, column = 2).value
        description = Drugbank.cell(row = i, column = 3).value
        interactions = Drugbank.cell(row = i, column = 4).value
        
        cur.execute("INSERT INTO `dtbank`.`Drugs` ( `id`, `name`, `description`) VALUES ( %s, %s, %s)",(id1,name,description))
        connection.commit()

        
        list1=interactions.strip('][').split(', ')
        if len(list1)==0 or list1[0]=='':
            continue
        list1=[id[1:-1] for id in list1]
        print(list1,'#')

        
        for i in range(0,len(list1)):
            try:
                print(id1,list1[i])
                cur.execute("INSERT INTO `dtbank`.`InteractsWith` ( `id1`, `id2`) VALUES ( %s, %s)",(id1,list1[i]))
                connection.commit()
            except:
                continue

    UniProt=data['UniProt']

    nofRows=UniProt.max_row
    

    for i in range(2, nofRows + 1):
        id = UniProt.cell(row = i, column = 1).value
        seq = UniProt.cell(row = i, column = 2).value
        
        
        cur.execute("INSERT INTO `dtbank`.`Uniprot` ( `id`, `seq`) VALUES ( %s, %s)",(id,seq))
        connection.commit()
        

    BindingDB=data['BindingDB']

    nofRows=BindingDB.max_row
    

    for i in range(2, nofRows + 1):
        reactionId = BindingDB.cell(row = i, column = 1).value
        drugbankId = BindingDB.cell(row = i, column = 2).value
        uniprotId = BindingDB.cell(row = i, column = 3).value
        target = BindingDB.cell(row = i, column = 4).value
        smiles = BindingDB.cell(row = i, column = 5).value
        measure = BindingDB.cell(row = i, column = 6).value
        affinity = BindingDB.cell(row = i, column = 7).value
        doi = BindingDB.cell(row = i, column = 8).value
        authors = BindingDB.cell(row = i, column = 9).value
        institute = BindingDB.cell(row = i, column = 10).value

        # print(reactionId,drugbankId,uniprotId,target,smiles,measure,affinity,doi,authors,institute)
        cur.execute("UPDATE `dtbank`.`Drugs` SET `smile` = %s WHERE (`id` = %s);",(smiles,drugbankId))
        connection.commit()
        cur.execute("UPDATE `dtbank`.`UniProt` SET `name` = %s WHERE (`id` = %s);",(target,uniprotId))
        connection.commit()
        try:
            cur.execute("INSERT INTO `dtbank`.`institution` ( `name`, `points`) VALUES ( %s, %s)",(institute,0))
        except:
            print(' ')

        try:
            cur.execute("INSERT INTO `dtbank`.`publications` ( `doi`, `institution`) VALUES ( %s, %s)",(doi,institute))
        except:
            print(' ')
        try:
            list1=authors.split('; ')
            
            for name in list1:
                print(name)
                try:
                    cur.execute("SELECT username FROM Users WHERE name=%s AND institution=%s",(name,institute))
                    username=cur.fetchone()
                    cur.execute("INSERT INTO `dtbank`.`contributors` ( `doi`,`author`, `institution`) VALUES ( %s, %s,%s)",(doi,username,institute))
                    connection.commit()
                except:
                    print('')
            cur.execute("INSERT INTO `dtbank`.`BindingDB` ( `reaction_id`, `drug_id`, `prot_id`,`affinity`,`measure`,`doi`) VALUES ( %s, %s,%s,%s,%s,%s)",(reactionId,drugbankId,uniprotId,affinity,measure,doi))
            connection.commit()
        except:
            print
            ('')
    SIDER=data['SIDER']

    nofRows=SIDER.max_row
    

    for i in range(2, nofRows + 1):
        cui = SIDER.cell(row = i, column = 1).value
        id = SIDER.cell(row = i, column = 2).value
        name = SIDER.cell(row = i, column = 3).value

        try:
            cur.execute("INSERT INTO `dtbank`.`SideEffect` ( `id`, `name`) VALUES ( %s, %s)",(cui,name))
            connection.commit()
            cur.execute("INSERT INTO `dtbank`.`SIDER` ( `drug_id`, `side_id`) VALUES ( %s, %s)",(id,cui))
            connection.commit()
        except:
            continue
        # print(cui,id,name)
    
   
